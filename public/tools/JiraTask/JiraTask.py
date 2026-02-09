from flask import Flask, jsonify, request, send_file
from flask_cors import CORS
import os
import csv
from io import StringIO, BytesIO
from datetime import datetime, timedelta, timezone
from pathlib import Path
import re
from typing import Optional
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


import requests
from requests.auth import HTTPBasicAuth
from dotenv import load_dotenv


BASE_DIR = Path(__file__).resolve().parents[3]
SCRIPT_DIR = Path(__file__).resolve().parent

# Carica prima la .env root, poi una .env locale nella cartella JiraTask (se presente)
load_dotenv(BASE_DIR / ".env")
load_dotenv(SCRIPT_DIR / ".env")

app = Flask(__name__)
CORS(app)
app.logger.setLevel("INFO")
import logging, sys

app.logger.handlers.clear()

handler = logging.StreamHandler(sys.stdout)
handler.setLevel(logging.INFO)
formatter = logging.Formatter("[%(asctime)s] %(levelname)s in %(module)s: %(message)s")
handler.setFormatter(formatter)

app.logger.addHandler(handler)
app.logger.setLevel(logging.INFO)
app.logger.propagate = False

logging.getLogger("werkzeug").setLevel(logging.INFO)


CLOUD_ID = os.getenv("JSM_CLOUD_ID")
JIRA_URL = os.getenv("JIRA_URL")
# Se JIRA_URL è impostato usiamo solo quello; altrimenti proviamo a derivare dallo JSM_SITE_URL
if not JIRA_URL:
    site = os.getenv("JSM_SITE_URL", "").rstrip("/")
    if site:
        JIRA_URL = f"{site}/rest/api/3"
    else:
        JIRA_URL = "https://cap4cloud.atlassian.net/rest/api/3"

# Fallback sugli env di JSM se non definiti quelli specifici Jira
JIRA_USERNAME = os.getenv("JIRA_USERNAME") or os.getenv("JSM_EMAIL")
JIRA_PASSWORD = os.getenv("JIRA_PASSWORD") or os.getenv("JSM_API_TOKEN")
BOT_USERS = [
    user.strip()
    for user in (os.getenv("JIRA_BOT_USERS") or "cap4bot").split(",")
    if user.strip()
]

CONFLUENCE_BASE = os.getenv("CONFLUENCE_BASE", "https://cap4cloud.atlassian.net/wiki").rstrip("/")
CONFLUENCE_EMAIL = os.getenv("CONFLUENCE_EMAIL")
CONFLUENCE_TOKEN = os.getenv("CONFLUENCE_TOKEN")

PROJECTS_ALL = [
    p.strip() for p in (os.getenv(
        "JIRA_PROJECTS_ALL",
        "IMSL1GEN,LVMH03MMS,LVMH01MMS,LVMH02MMS,LoroPiana,HBLT01IMS,WTJW01IMS"
    )).split(",") if p.strip()
]
app.logger.info(f"[JiraTask] PROJECTS_ALL={PROJECTS_ALL}")


if not JIRA_USERNAME or not JIRA_PASSWORD:
    raise Exception("JIRA credentials not set in environment variables")

DEFAULT_BASES = [JIRA_URL.rstrip("/")]
# Se vuoi forzare anche l'endpoint api.atlassian.com (OAuth/Bearer), abilita JIRA_ALLOW_EX_API=true
if os.getenv("JIRA_ALLOW_EX_API", "false").lower() == "true" and CLOUD_ID:
    DEFAULT_BASES.append(f"https://api.atlassian.com/ex/jira/{CLOUD_ID}/rest/api/3")


def build_base_urls():
    seen = set()
    ordered = []
    for base in DEFAULT_BASES:
        if base and base not in seen:
            ordered.append(base)
            seen.add(base)
    return ordered


BASE_URLS = build_base_urls()
app.logger.info(f"[JiraTask] BASE_URLS={BASE_URLS} USER={JIRA_USERNAME}")


def parse_jira_datetime(raw: str) -> datetime:
    """Normalizza i datetime Jira (+0000/+00:00/Z) in datetime Python."""
    cleaned = (raw or "").replace("Z", "+00:00")
    match = re.search(r"([+-]\d{2})(\d{2})$", cleaned)
    if match:
        cleaned = cleaned[:-5] + f"{match.group(1)}:{match.group(2)}"
    return datetime.fromisoformat(cleaned)


def parse_time_window(
    start_date: Optional[str],
    end_date: Optional[str],
    shift_start: Optional[str],
    shift_end: Optional[str],
):
    """Crea un intervallo datetime con timezone per il filtro. Se mancano filtri usa tutto il giorno corrente."""
    today_str = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    start_date = (start_date or today_str).strip()
    end_date = (end_date or start_date).strip()
    shift_start = shift_start or "00:00"
    shift_end = shift_end or "23:59"

    if not start_date:
        raise ValueError("Il parametro startDate è obbligatorio.")

    fmt = "%Y-%m-%d %H:%M"
    start_dt = datetime.strptime(f"{start_date} {shift_start}", fmt).replace(tzinfo=timezone.utc)
    end_dt = datetime.strptime(f"{end_date} {shift_end}", fmt).replace(tzinfo=timezone.utc)

    # Gestisce turni che oltrepassano la mezzanotte (es. 22:00 -> 06:00)
    if end_dt <= start_dt:
        end_dt += timedelta(days=1)

    return start_dt, end_dt


def build_jql_query(start_dt, end_dt, project, agent_account_id=None) -> str:
    start_str = start_dt.strftime("%Y-%m-%d %H:%M")
    end_str = (end_dt + timedelta(minutes=1)).strftime("%Y-%m-%d %H:%M")
    base_time = f"updated >= '{start_str}' AND updated < '{end_str}'"

    proj = (project or "").strip()
    if not proj or proj.lower() == "all":
        projects = ",".join([f'"{p}"' for p in PROJECTS_ALL])
        scope = f"project in ({projects})"
    else:
        scope = f'project = "{proj}"'

    agent_clause = ""
    if agent_account_id:
        aid = f'accountId("{agent_account_id}")'
        agent_clause = (
            " AND ("
            f"assignee = {aid} "
            f"OR assignee WAS {aid} "
            f"OR reporter = {aid}"
            ")"
        )

    return f"{scope} AND {base_time}{agent_clause}"




import time
import requests
DEFAULT_TIMEOUT = (5, 60)  # connect, read


def jira_request(method: str, path: str, params=None, json=None):
    last_err = None
    max_retries_429 = 5

    for base in BASE_URLS:
        url = f"{base.rstrip('/')}/{path.lstrip('/')}"
        for attempt in range(max_retries_429 + 1):
            try:
                resp = requests.request(
                    method=method.upper(),
                    url=url,
                    params=params,
                    json=json,
                    auth=HTTPBasicAuth(JIRA_USERNAME, JIRA_PASSWORD),
                    headers={"Accept": "application/json", "User-Agent": "JiraTask/1.0"},
                    timeout=(5, 60),
                )

                if resp.status_code == 429:
                    retry_after = resp.headers.get("Retry-After")
                    wait = int(retry_after) if retry_after and retry_after.isdigit() else (2 ** attempt)
                    app.logger.warning("[JiraTask] 429 rate limit %s wait=%ss", url, wait)
                    time.sleep(wait)
                    continue

                if not resp.ok:
                    app.logger.warning("[JiraTask] Jira HTTP %s %s body=%s", resp.status_code, url, (resp.text or "")[:800])

                resp.raise_for_status()
                return resp

            except requests.HTTPError as err:
                last_err = err
                status = err.response.status_code if err.response else "?"
                if status in (401, 403):
                    raise
                break
            except requests.RequestException as err:
                last_err = err
                break

    if last_err:
        raise last_err
    raise RuntimeError("Nessun endpoint Jira raggiungibile")


def confluence_request(path, params=None):
    url = f"{CONFLUENCE_BASE}{path}"
    resp = requests.get(
        url,
        params=params,
        auth=HTTPBasicAuth(CONFLUENCE_EMAIL, CONFLUENCE_TOKEN),
        headers={"Accept": "application/json"},
        timeout=(5, 60),
    )
    resp.raise_for_status()
    return resp


def _search_jql_new(jql_query: str, max_results: int, next_page_token: str | None = None):
    payload = {
        "jql": jql_query,
        "maxResults": max_results,
        "fields": ["key", "summary", "status", "assignee", "updated", "project"],
    }
    if next_page_token:
        payload["nextPageToken"] = next_page_token

    resp = jira_request("post", "search/jql", json=payload)
    data = resp.json() or {}

    issues = data.get("issues") or []
    token = data.get("nextPageToken")

    app.logger.info("[JiraTask] /search/jql issues=%s nextToken=%s", len(issues), "yes" if token else "no")
    return issues, token, data.get("total")





_AGENT_CACHE = {}  # displayName -> {"accountId":..., "displayName":...}

def resolve_agent_account_id(agent_name: str) -> Optional[str]:
    """
    Converte "Jean Molina" -> accountId Jira usando /user/search.
    Cache in memoria per non chiamare Jira ogni volta.
    """
    name = (agent_name or "").strip()
    if not name or name.lower() == "all":
        return None

    if name.lower() in _AGENT_CACHE:
        return _AGENT_CACHE[name.lower()]["accountId"]

    resp = jira_request("get", "user/search", params={"query": name, "maxResults": 50})
    users = resp.json() or []

    def norm(s: str) -> str:
        return " ".join((s or "").lower().split())

    target = norm(name)
    best = None

    # prova match esatto su displayName
    for u in users:
        if norm(u.get("displayName", "")) == target:
            best = u
            break

    # fallback: primo risultato
    if not best and users:
        best = users[0]

    if not best:
        app.logger.warning(f"[JiraTask] Agent '{name}' not found via user/search")
        return None

    account_id = best.get("accountId")
    _AGENT_CACHE[name.lower()] = {"accountId": account_id, "displayName": best.get("displayName")}
    app.logger.info(f"[JiraTask] Resolved agent '{name}' -> accountId={account_id} displayName='{best.get('displayName')}'")
    return account_id



def _approximate_count(jql_query: str) -> int | None:
    """Chiama /search/approximate-count per capire se la JQL produce match lato Jira."""
    try:
        resp = jira_request("post", "search/approximate-count", json={"jql": jql_query})
        data = resp.json() or {}
        count = data.get("count")
        app.logger.info("[JiraTask] approximate-count for JQL: %s -> %s", jql_query, count)
        return count
    except requests.HTTPError as err:
        body = ""
        try:
            body = (err.response.text or "").strip()
        except Exception:
            body = ""
        app.logger.warning(
            "[JiraTask] approximate-count HTTP %s body='%s'",
            err.response.status_code if err.response else "?",
            body[:400],
        )
    except Exception as err:
        app.logger.warning(f"[JiraTask] approximate-count errore: {err}")
    return None

def fetch_issues(jql_query: str, max_issues: Optional[int] = None):
    issues = []
    max_results = 100
    next_token = None
    page = 0

    if max_issues is None:
        max_issues = int(os.getenv("JIRA_MAX_ISSUES", "200"))

    while True:
        batch, next_token, _ = _search_jql_new(jql_query, max_results, next_token)
        page += 1
        app.logger.info(
            "[JiraTask] page=%s batch=%s nextToken=%s",
            page,
            len(batch),
            "yes" if next_token else "no",
        )

        issues.extend(batch)

        if len(issues) >= max_issues:
            app.logger.info("[JiraTask] raggiunto limite max_issues=%s, stop", max_issues)
            break

        if not next_token:
            break

    return issues




def fetch_issue_changelog(issue_key: str):
    """
    Recupera il changelog di una singola issue.
    Gestisce il 429 (rate limit) tornando un changelog vuoto
    invece di far esplodere tutta la richiesta.
    """
    try:
        response = jira_request("get", f"issue/{issue_key}/changelog")
        return response.json() or {}
    except requests.HTTPError as err:
        status = err.response.status_code if err.response else None
        body = ""
        try:
            body = (err.response.text or "").strip()
        except Exception:
            body = ""

        if status == 429:
            app.logger.warning(
                f"[JiraTask] changelog {issue_key} rate-limited (429), skip. body='{body[:200]}'"
            )
            # Niente eccezione: issue senza attività
            return {"values": []}

        app.logger.warning(
            f"[JiraTask] changelog fetch error for {issue_key}: HTTP {status} body='{body[:200]}'"
        )
        # Anche qui non rilancio: considero semplicemente questa issue senza attività
        return {"values": []}
    

import html
from html.parser import HTMLParser

class _MLStripper(HTMLParser):
    def __init__(self):
        super().__init__()
        self.parts = []
    def handle_data(self, d):
        self.parts.append(d)

def html_to_text(s: str) -> str:
    if not s:
        return ""
    stripper = _MLStripper()
    stripper.feed(s)
    return html.unescape("".join(stripper.parts)).strip()

def adf_to_text(node) -> str:
    """
    Converte Atlassian Document Format (ADF) -> testo semplice.
    Non perfetto ma rende molto bene per commenti normali.
    """
    if node is None:
        return ""
    if isinstance(node, str):
        return node
    if isinstance(node, list):
        return "".join(adf_to_text(n) for n in node)

    t = node.get("type")
    if t == "text":
        return node.get("text", "")
    if t in ("paragraph", "heading", "blockquote"):
        return adf_to_text(node.get("content", [])) + "\n"
    if t == "hardBreak":
        return "\n"
    if t in ("bulletList", "orderedList"):
        return adf_to_text(node.get("content", [])) + "\n"
    if t == "listItem":
        return "- " + adf_to_text(node.get("content", [])) + "\n"
    if t == "codeBlock":
        return adf_to_text(node.get("content", [])) + "\n"
    # fallback generico
    return adf_to_text(node.get("content", []))

def fetch_issue_comments_for_agent(issue_key: str, agent_account_id: str, start_dt: datetime, end_dt: datetime):
    """
    Prende SOLO i commenti dell'agente selezionato, con body completo.
    Paginazione supportata.
    """
    if not agent_account_id:
        return []

    out = []
    start_at = 0
    max_results = 100

    while True:
        try:
            resp = jira_request(
                "get",
                f"issue/{issue_key}/comment",
                params={"startAt": start_at, "maxResults": max_results, "expand": "renderedBody"},
            )
        except requests.HTTPError as err:
            status = err.response.status_code if err.response else None
            if status == 429:
                app.logger.warning("[JiraTask] comments %s rate-limited (429), skip", issue_key)
                return []
            app.logger.warning("[JiraTask] comments fetch error %s HTTP %s", issue_key, status)
            return []
        data = resp.json() or {}
        comments = data.get("comments", []) or []

        for c in comments:
            author = (c.get("author") or {})
            if author.get("accountId") != agent_account_id:
                continue

            created_raw = c.get("created")
            if not created_raw:
                continue

            try:
                created_dt = parse_jira_datetime(created_raw)
                if created_dt.tzinfo is None:
                    created_dt = created_dt.replace(tzinfo=timezone.utc)
            except Exception:
                continue

            if not (start_dt <= created_dt <= end_dt):
                continue

            if _is_in_handover_window(created_dt):
                continue

            # body completo: preferisco renderedBody (HTML), altrimenti ADF
            body = ""
            if c.get("renderedBody"):
                body = html_to_text(c.get("renderedBody"))
            else:
                body_obj = c.get("body") or {}
                # Jira Cloud comment body spesso è ADF: {"type":"doc","content":[...]}
                body = adf_to_text(body_obj)

            out.append({
                "field": "comment",
                "from": "",
                "to": body,
                "author": author.get("displayName", ""),
                "date": created_raw,
            })

        start_at += len(comments)
        if start_at >= int(data.get("total", len(comments))):
            break
        if not comments:
            break

    return out




# --- HANDOVER & ATTIVITÀ ----------------------------------------------------

HANDOVER_TIMES = [(6, 0), (14, 0), (22, 0)]  # 6:00, 14:00, 22:00
HANDOVER_BEFORE_MIN = 30                    # mezz'ora prima
HANDOVER_AFTER_MIN = 60                     # un'ora dopo


def _is_in_handover_window(dt: datetime) -> bool:
    """
    Ritorna True se dt cade entro una finestra di mezz'ora prima / un'ora dopo
    uno degli orari di handover (6:00, 14:00, 22:00).
    """
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)

    for h, m in HANDOVER_TIMES:
        handover = dt.replace(hour=h, minute=m, second=0, microsecond=0)
        start = handover - timedelta(minutes=HANDOVER_BEFORE_MIN)
        end = handover + timedelta(minutes=HANDOVER_AFTER_MIN)
        if start <= dt <= end:
            return True
    return False


def filter_issue_activities(
    issue_key,
    changelog,
    agent_name,
    agent_account_id,
    bot_users,
    start_dt,
    end_dt,
):
    """
    Filtra attività da un changelog Jira, ignorando:
      - attività dei bot
      - attività nei periodi di handover (±30m / +60m da 6:00, 14:00, 22:00)
      - history che modificano SOLO l'assignee
    e tenendo solo quelle dell'agente selezionato (se diverso da All).
    """
    if not changelog:
        return []

    histories = (
        changelog.get("histories")
        or changelog.get("values")
        or []
    )

    relevant = []
    bot_set = {b.lower() for b in bot_users}

    only_bot = True

    for history in histories:
        author_obj = history.get("author", {}) or {}
        author_name = author_obj.get("displayName", "")
        author_account_id = author_obj.get("accountId")
        author_norm = " ".join(author_name.lower().split())

        # filtro agente (se selezionato)
        if agent_account_id and author_account_id != agent_account_id:
            continue

        created_raw = history.get("created")
        if not created_raw:
            continue

        try:
            created_date = parse_jira_datetime(created_raw)
            if created_date.tzinfo is None:
                created_date = created_date.replace(tzinfo=timezone.utc)
        except Exception:
            app.logger.warning(
                f"[JiraTask] invalid history date in {issue_key}: {created_raw}"
            )
            continue

        # Finestra temporale generale
        if not (start_dt <= created_date <= end_dt):
            continue

        # Ignora le attività dentro le finestre di handover
        if _is_in_handover_window(created_date):
            app.logger.debug(
                f"[JiraTask] history in handover window for {issue_key} at {created_date}, skip"
            )
            continue

        # Autore bot?
        if author_norm not in bot_set:
            only_bot = False
        else:
            # Attività di bot: non contano
            continue

        items = history.get("items", []) or []

        # Se TUTTI i campi modificati sono "assignee", ignora (handover massivo / reassignment pura)
        if items and all((item.get("field") or "").lower() == "assignee" for item in items):
            app.logger.debug(f"[JiraTask] history only-assignee for {issue_key}, skip")
            continue

        for item in items:
            relevant.append({
                "field": item.get("field"),
                "from": item.get("fromString"),
                "to": item.get("toString"),
                "author": author_name,
                "date": created_raw,
            })

    if only_bot:
        return []

    return relevant



def build_task_payload(issue, activities):
    fields = issue.get("fields") or {}

    updated_raw = fields.get("updated", "")
    updated = updated_raw
    if updated_raw:
        try:
            updated = parse_jira_datetime(updated_raw).strftime("%d/%m/%Y %H:%M:%S")
        except ValueError:
            app.logger.warning(
                f"Formato data non valido in updated per {issue.get('key', 'UNKNOWN')}: {updated_raw}"
            )

    assignee_field = fields.get("assignee") or {}
    assignee_name = assignee_field.get("displayName", "Unassigned")

    status_field = fields.get("status") or {}
    status_name = status_field.get("name", "Unknown")

    return {
        "task_name": issue.get("key", ""),
        "task_link": f"https://cap4cloud.atlassian.net/browse/{issue.get('key', '')}",
        "summary": fields.get("summary", ""),
        "task_status": status_name,
        "assignee": assignee_name,
        "updated": updated,
        "activities": activities,
    }


def collect_task_results(start_date, end_date, project, shift_start, shift_end, agent):
    start_dt, end_dt = parse_time_window(start_date, end_date, shift_start, shift_end)

    agent_account_id = resolve_agent_account_id(agent)
    agent_is_all = (not agent) or (agent.lower() == "all")

    jql_query = build_jql_query(start_dt, end_dt, project, agent_account_id)
    app.logger.info(f"[JiraTask] JQL={jql_query}")

    span_days = (end_dt.date() - start_dt.date()).days + 1
    max_approx = int(os.getenv("JIRA_MAX_APPROX_RESULTS", "800"))

    max_issues = int(os.getenv("JIRA_MAX_ISSUES", "200"))
    if not agent_is_all:
        max_issues = int(os.getenv("JIRA_MAX_ISSUES_AGENT", "2000"))

    if span_days > 3 and agent_is_all:
        approx = _approximate_count(jql_query)
        if approx is not None and approx > max_approx:
            raise ValueError(
                f"La query Jira restituirebbe circa {approx} issue, oltre il limite consentito ({max_approx}). "
                "Riduci l'intervallo o filtra meglio."
            )

    issues = fetch_issues(jql_query, max_issues=max_issues)
    
    from collections import Counter

    issues = fetch_issues(jql_query, max_issues=max_issues)

    by_project = Counter()
    for i in issues:
        proj = ((i.get("fields") or {}).get("project") or {}).get("key") or "UNKNOWN"
        by_project[proj] += 1

    app.logger.info("[JiraTask] Issues by project: %s", dict(by_project))

    app.logger.info(f"[JiraTask] Issues fetched={len(issues)} (max_issues={max_issues})")

    def process_issue(issue):
        key = issue.get("key", "UNKNOWN")

        changelog = fetch_issue_changelog(key)
        activities = filter_issue_activities(
            key, changelog, agent, agent_account_id, BOT_USERS, start_dt, end_dt
        )

        comments = []
        if agent_account_id:
            comments = fetch_issue_comments_for_agent(key, agent_account_id, start_dt, end_dt)

        merged = (activities or []) + (comments or [])
        if not merged:
            return None

        def _dt(a):
            try:
                return parse_jira_datetime(a.get("date", ""))
            except Exception:
                return datetime.max.replace(tzinfo=timezone.utc)

        merged.sort(key=_dt)
        return build_task_payload(issue, merged)

    results = []
    max_workers = int(os.getenv("JIRA_CHANGELOG_WORKERS", "5"))
    app.logger.info(f"[JiraTask] processing issues with ThreadPoolExecutor(max_workers={max_workers})")

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = [executor.submit(process_issue, issue) for issue in issues]
        for f in as_completed(futures):
            payload = f.result()
            if payload:
                results.append(payload)

    app.logger.info(f"[JiraTask] Final results count={len(results)}")
    return results, jql_query




# --- HTTP UTILS & ENDPOINTS -------------------------------------------------


@app.route("/confluence/page-comments", methods=["GET"])
def confluence_page_comments():
    page_id = request.args.get("pageId")
    author = request.args.get("author")  # displayName (best-effort)
    start_date = request.args.get("startDate")
    end_date = request.args.get("endDate", start_date)

    start_dt, end_dt = parse_time_window(start_date, end_date, "00:00", "23:59")

    # v1 endpoint: comment children
    resp = confluence_request(
        f"/rest/api/content/{page_id}/child/comment",
        params={"expand": "results.history,results.body.view", "limit": 200},
    )
    data = resp.json() or {}
    results = data.get("results") or []

    out = []
    for c in results:
        hist = (c.get("history") or {}).get("createdBy") or {}
        created = (c.get("history") or {}).get("createdDate")
        if not created:
            continue

        created_dt = parse_jira_datetime(created)  # funziona anche qui se è ISO
        if not (start_dt <= created_dt <= end_dt):
            continue

        if author and (hist.get("displayName") != author):
            continue

        body_html = (((c.get("body") or {}).get("view") or {}).get("value")) or ""
        out.append({
            "type": "comment",
            "pageId": page_id,
            "author": hist.get("displayName"),
            "date": created,
            "text": html_to_text(body_html),
        })

    return jsonify(results=out, total=len(out))


def _common_params():
    start_date = request.args.get("startDate")
    end_date = request.args.get("endDate", start_date)
    project = request.args.get("project", "")
    shift_start = request.args.get("shiftStart", "00:00")
    shift_end = request.args.get("shiftEnd", "23:59")
    agent = request.args.get("agent", "All")
    return start_date, end_date, project, shift_start, shift_end, agent


def _http_error_message(err: requests.HTTPError) -> str:
    if not err.response:
        return str(err)
    try:
        data = err.response.json()
        if isinstance(data, dict):
            # Jira tipicamente usa errorMessages[]
            msgs = data.get("errorMessages")
            if msgs:
                return "; ".join(msgs)
            if "message" in data:
                return str(data["message"])
    except Exception:
        pass
    return err.response.text or str(err)


@app.route("/check-tasks", methods=["GET"])
def check_tasks():
    app.logger.info("[JiraTask] HIT /check-tasks args=%s", dict(request.args))
    try:
        params = _common_params()
        results, jql_query = collect_task_results(*params)
        app.logger.info(f"JQL: {jql_query} - risultati: {len(results)}")
        return jsonify(results=results, total_analyzed=len(results))
    except ValueError as e:
        return jsonify(error=str(e)), 400
    except requests.HTTPError as e:
        status = e.response.status_code if e.response else 502
        body = _http_error_message(e)
        app.logger.error(f"HTTPError Jira ({status}) {e.request.url if e.request else ''} -> {body}")
        return jsonify(error=f"Errore Jira ({status}): {_http_error_message(e)}"), status
    except Exception as e:
        app.logger.exception(f"Errore imprevisto: {e}")
        return jsonify(error="Errore interno, controlla i log."), 500


@app.route("/export-xlsx", methods=["GET"])
def export_xlsx():
    try:
        params = _common_params()
        results, jql_query = collect_task_results(*params)

        if not results:
            return jsonify({"error": "Nessun dato trovato per esportare."}), 404

        # flatten
        flat_rows = []
        for item in results:
            for act in item.get("activities", []):
                flat_rows.append([
                    clean_xlsx_value(item.get("task_name", "")),
                    clean_xlsx_value(item.get("task_link", "")),
                    clean_xlsx_value(item.get("summary", "")),
                    clean_xlsx_value(act.get("date", "")),
                    clean_xlsx_value(act.get("author", "")),
                    clean_xlsx_value(act.get("field", "")),
                    clean_xlsx_value(act.get("from", "")),
                    clean_xlsx_value(act.get("to", "")),
                ])


        # ordina per data attività
        def _dt(v):
            try:
                return parse_jira_datetime(v)
            except Exception:
                return datetime.max.replace(tzinfo=timezone.utc)

        flat_rows.sort(key=lambda r: _dt(r[3]))

        wb = Workbook()
        ws = wb.active
        ws.title = "Timeline"

        headers = ["Task Key","Task Link","Summary","Activity Date","Author","Field","From","To/Comment"]
        ws.append(headers)

        wrap = Alignment(wrap_text=True, vertical="top")
        top = Alignment(vertical="top")

        for r in flat_rows:
            ws.append(r)

        # stile: wrap su Summary e To/Comment
        summary_col = headers.index("Summary") + 1
        to_col = headers.index("To/Comment") + 1

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = top
            row[summary_col-1].alignment = wrap
            row[to_col-1].alignment = wrap

        # auto width (con max)
        max_width = 80
        for col_idx, col in enumerate(ws.columns, start=1):
            max_len = 0
            for cell in col:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            width = min(max_len + 2, max_width)
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # salva in memoria
        out = BytesIO()
        wb.save(out)
        out.seek(0)

        app.logger.info(f"XLSX esportato per JQL: {jql_query} ({len(flat_rows)} attività)")
        return send_file(
            out,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name="tasks_timeline.xlsx",
        )

    except Exception as e:
        app.logger.exception("[JiraTask] export-xlsx crashed: %s", e)
        return jsonify({"error": "Errore interno, controlla i log."}), 500


import re
_ILLEGAL_XLSX = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")

def clean_xlsx_value(v):
    if v is None:
        return ""
    s = str(v)
    return _ILLEGAL_XLSX.sub("", s)


@app.route("/", methods=["GET"])
@app.route("/health", methods=["GET"])
def health():
    return jsonify({"status": "ok", "base_urls": BASE_URLS}), 200


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=2000, debug=True, use_reloader=False)


@app.route("/debug/projects", methods=["GET"])
def debug_projects():
    resp = jira_request("get", "project/search", params={"maxResults": 1000})
    data = resp.json() or {}
    values = data.get("values") or data.get("projects") or []
    projects = [{"key": p.get("key"), "name": p.get("name")} for p in values]
    return jsonify(projects=projects)